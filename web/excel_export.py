"""Export statistics to an .xlsx workbook using openpyxl."""
from __future__ import annotations

import io
import datetime as dt
from common import time_utils

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from common.models import Order, User


def _header(ws, headers: list[str]) -> None:
    fill = PatternFill("solid", fgColor="0D6EFD")
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill


def build_stats_workbook(session: Session) -> io.BytesIO:
    wb = Workbook()

    # --- Orders sheet ---
    ws = wb.active
    ws.title = "Заказы"
    _header(ws, ["ID", "Статус", "Откуда", "Куда", "Цена", "Ожидание", "Км", "Создан"])
    for o in session.query(Order).order_by(Order.created_at.desc()).all():
        ws.append([
            o.id, o.status, o.address_from, o.address_to,
            float(o.price or 0), float(o.waiting_fee or 0),
            o.distance_km, time_utils.format_local(o.created_at, "%Y-%m-%d %H:%M") if o.created_at else "",
        ])

    # --- Drivers sheet ---
    ws2 = wb.create_sheet("Водители")
    # Requirement 10: the «Заработано» column was removed together with the
    # ``total_earned`` DB field. Show rating + number of reviews instead.
    _header(ws2, ["ID", "Имя", "VK ID", "Авто", "Рейтинг", "Отзывов"])
    for d in session.query(User).filter(User.role == "driver").all():
        ws2.append([
            d.id, d.full_name, d.vk_id,
            f"{d.car_model or ''} {d.car_number or ''}".strip(),
            d.rating, int(d.rating_count or 0),
        ])

    # --- Summary sheet ---
    ws3 = wb.create_sheet("Сводка")
    total_orders = session.query(Order).count()
    completed = session.query(Order).filter(Order.status == "completed").count()
    revenue = sum(float(o.price or 0) for o in session.query(Order).filter(Order.status == "completed"))
    ws3.append(["Метрика", "Значение"])
    ws3.append(["Всего заказов", total_orders])
    ws3.append(["Завершено", completed])
    ws3.append(["Выручка, ₽", round(revenue, 2)])
    ws3.append(["Сформировано", time_utils.now().strftime("%Y-%m-%d %H:%M")])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


ORDER_STATUS_LABELS = {
    "created": "Новая",
    "queued": "В очереди",
    "searching": "Ожидает водителя",
    "chat_search": "Поиск в чате",
    "assigned": "Водитель назначен",
    "arrived": "Водитель прибыл",
    "in_progress": "В пути",
    "completed": "Выполнен",
    "cancelled": "Отменён",
    "no_drivers": "Нет водителей",
    "parallel_assigned": "Параллельная заявка",
}


def _local_datetime(value) -> str:
    return time_utils.format_local(value, "%d.%m.%Y %H:%M") if value else ""


def _money(value) -> float | None:
    return float(value) if value is not None else None


def build_orders_workbook(orders: list[Order]) -> io.BytesIO:
    """Build a polished workbook from the exact orders selected in the UI."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Заказы"
    headers = [
        "№", "Создан", "Статус", "Тип", "Источник",
        "Пассажир", "Телефон пассажира", "VK пассажира",
        "Водитель", "VK водителя", "Откуда", "Куда", "Полный маршрут",
        "Комментарий", "Цена, ₽", "Платное ожидание, ₽", "Скидка, ₽",
        "Промокод", "Расстояние, км", "Время в пути, мин", "Линия",
        "Город подачи", "ETA, мин", "Принят водителем", "Водитель прибыл",
        "Завершён", "Отменён", "Кем отменён", "Количество отказов", "Оценка",
    ]
    _header(ws, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(headers)).coordinate}"
    ws.row_dimensions[1].height = 30

    for order in orders:
        passenger = order.passenger
        driver = order.driver
        source = (
            f"Диспетчер #{order.dispatcher_id}"
            if order.dispatcher_id else "Пассажир"
        )
        passenger_name = order.customer_name or (
            passenger.full_name if passenger else ""
        )
        passenger_phone = order.customer_phone or (
            passenger.phone if passenger else ""
        )
        route = order.route_text or " — ".join(
            value for value in (order.address_from, order.address_to) if value
        )
        ws.append([
            order.id,
            _local_datetime(order.created_at),
            ORDER_STATUS_LABELS.get(order.status, order.status or ""),
            "Доставка" if order.order_type == "delivery" else "Поездка",
            source,
            passenger_name,
            passenger_phone,
            passenger.vk_id if passenger else None,
            driver.full_name if driver else "",
            driver.vk_id if driver else None,
            order.address_from,
            order.address_to,
            route,
            order.comment or "",
            _money(order.price),
            _money(order.waiting_fee),
            _money(order.discount),
            order.promocode or "",
            order.distance_km,
            order.duration_min,
            order.line or "",
            order.pickup_city or "",
            order.arrival_eta,
            _local_datetime(order.driver_accept_time),
            _local_datetime(order.arrived_at),
            _local_datetime(order.completed_at),
            _local_datetime(order.cancelled_at),
            order.cancelled_by or "",
            int(order.decline_count or 0),
            order.rating,
        ])

    # Human-friendly widths with a safe cap for long routes/comments.
    widths = {
        1: 9, 2: 18, 3: 22, 4: 13, 5: 18, 6: 28, 7: 20, 8: 15,
        9: 28, 10: 15, 11: 28, 12: 28, 13: 50, 14: 38,
        15: 14, 16: 20, 17: 13, 18: 14, 19: 17, 20: 20, 21: 18,
        22: 18, 23: 12, 24: 20, 25: 20, 26: 20, 27: 20, 28: 18,
        29: 20, 30: 12,
    }
    for index, width in widths.items():
        ws.column_dimensions[ws.cell(row=1, column=index).column_letter].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column in (11, 12, 13, 14))
    for row in ws.iter_rows(min_row=2, min_col=15, max_col=17):
        for cell in row:
            cell.number_format = '#,##0.00'

    info = wb.create_sheet("Информация")
    info.append(["Параметр", "Значение"])
    info.append(["Количество заказов", len(orders)])
    info.append(["Сформировано", time_utils.format_local(time_utils.now(), "%d.%m.%Y %H:%M")])
    _header(info, ["Параметр", "Значение"])
    info.column_dimensions["A"].width = 25
    info.column_dimensions["B"].width = 30

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
