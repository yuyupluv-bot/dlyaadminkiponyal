from __future__ import annotations

import ast
import datetime as dt
import importlib.util
import pathlib
import sys
import types
import unittest
from decimal import Decimal
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import load_workbook

ROOT = pathlib.Path(__file__).resolve().parents[1]
APP = (ROOT / "web/app.py").read_text("utf-8")
EXPORT = (ROOT / "web/excel_export.py").read_text("utf-8")
TEMPLATE = (ROOT / "web/templates/orders.html").read_text("utf-8")


def function_source(source: str, name: str) -> str:
    tree = ast.parse(source)
    for node in tree.body:
        if isinstance(node, ast.FunctionDef) and node.name == name:
            return ast.get_source_segment(source, node) or ""
    raise AssertionError(name)


class OrdersExcelExportV92(unittest.TestCase):
    def test_01_sources_parse_and_bot_was_not_given_export_route(self) -> None:
        ast.parse(APP)
        ast.parse(EXPORT)
        self.assertNotIn("\ufffd", APP + EXPORT + TEMPLATE)
        self.assertNotIn('/orders/export.xlsx', (ROOT.parent / 'bot/bot/handlers.py').read_text('utf-8'))

    def test_02_export_route_is_authenticated_and_unlimited(self) -> None:
        self.assertIn('@app.route("/orders/export.xlsx")\n@login_required\ndef export_orders():', APP)
        body = function_source(APP, "export_orders")
        self.assertIn("_orders_query(db(), status, q)", body)
        self.assertIn("build_orders_workbook(rows)", body)
        self.assertIn('"export_orders"', body)
        self.assertNotIn(".limit(200)", body)
        self.assertIn("send_file", body)

    def test_03_screen_and_export_share_identical_filters(self) -> None:
        query = function_source(APP, "_orders_query")
        self.assertIn("Order.status == status", query)
        self.assertIn("Order.passenger.has(passenger_name)", query)
        self.assertIn("Order.driver.has(driver_name)", query)
        self.assertIn("Order.customer_name.ilike", query)
        screen = function_source(APP, "orders")
        self.assertIn("_orders_query(s, status, q)", screen)
        self.assertIn(".limit(200)", screen)

    def test_04_template_has_filtered_and_all_exports(self) -> None:
        self.assertIn("url_for('export_orders', status=status, q=q)", TEMPLATE)
        self.assertIn("url_for('export_orders')", TEMPLATE)
        self.assertIn("Выгрузить найденные", TEMPLATE)
        self.assertIn("Все заказы", TEMPLATE)
        self.assertIn("не только 200 строк", TEMPLATE)

    def test_05_workbook_contains_detailed_order_columns(self) -> None:
        fake_sqlalchemy = types.ModuleType("sqlalchemy")
        fake_orm = types.ModuleType("sqlalchemy.orm")
        fake_orm.Session = object
        fake_models = types.ModuleType("common.models")
        fake_models.Order = object
        fake_models.User = object
        spec = importlib.util.spec_from_file_location(
            "excel_export_v92_test", ROOT / "web/excel_export.py"
        )
        module = importlib.util.module_from_spec(spec)
        with patch.dict(sys.modules, {
            "sqlalchemy": fake_sqlalchemy,
            "sqlalchemy.orm": fake_orm,
            "common.models": fake_models,
        }):
            assert spec.loader is not None
            spec.loader.exec_module(module)
        build_orders_workbook = module.build_orders_workbook

        passenger = SimpleNamespace(full_name="Иван Иванов", phone="79990000000", vk_id=101)
        driver = SimpleNamespace(full_name="Водитель Тест", vk_id=202)
        now = dt.datetime(2026, 8, 14, 12, 0, tzinfo=dt.timezone.utc)
        order = SimpleNamespace(
            id=77, passenger=passenger, driver=driver, dispatcher_id=None,
            customer_name=None, customer_phone=None, created_at=now,
            status="completed", order_type="regular", address_from="А",
            address_to="Б", route_text="А — Б", comment="Комментарий",
            price=Decimal("350.00"), waiting_fee=Decimal("25.00"),
            discount=Decimal("10.00"), promocode="TEST", distance_km=5.5,
            duration_min=12, line="Горнозаводск", pickup_city="Горнозаводск",
            arrival_eta=7, driver_accept_time=now, arrived_at=now,
            completed_at=now, cancelled_at=None, cancelled_by=None,
            decline_count=1, rating=5,
        )
        book = load_workbook(build_orders_workbook([order]))
        self.assertEqual(book.sheetnames, ["Заказы", "Информация"])
        sheet = book["Заказы"]
        headers = [cell.value for cell in sheet[1]]
        for expected in (
            "Пассажир", "Телефон пассажира", "Водитель", "Полный маршрут",
            "Цена, ₽", "Платное ожидание, ₽", "Кем отменён", "Оценка",
        ):
            self.assertIn(expected, headers)
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertEqual(sheet.cell(2, 1).value, 77)
        self.assertEqual(sheet.cell(2, 6).value, "Иван Иванов")
        self.assertEqual(sheet.cell(2, 9).value, "Водитель Тест")
        self.assertEqual(sheet.cell(2, 15).value, 350.0)
        self.assertEqual(book["Информация"][2][1].value, 1)


if __name__ == "__main__":
    unittest.main()
